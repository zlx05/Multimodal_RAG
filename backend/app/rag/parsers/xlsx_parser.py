"""Excel（.xlsx）解析器：逐 sheet 展开合并单元格，按行组切成 table 块。

策略：
- 合并单元格先把左上角值广播到整个区域，避免大量空 cell 伤向量质量。
- 首个有内容行视为表头；sheet 名作上下文（heading_path=[sheet, 行 N-M]）。
- 每个行组块序列化成「列名：值；…」，块内重复表头保证自包含。
- 空 sheet 跳过；content_type="table"（屏障块，不被语义切分切碎）。
"""

from pathlib import Path

from .base import BaseParser
from .tabular import iter_row_groups, rows_to_nested, table_text
from ..blocks import DocumentBlock


class XlsxParser(BaseParser):
    source_type = "xlsx"

    def __init__(self, document_id: str, **kwargs):
        # original_dir/work_dir/vision 等由 worker 统一传入，xlsx 直接读原文件无需复制
        super().__init__(document_id)

    def parse(self, path: str | Path) -> list[DocumentBlock]:
        from openpyxl import load_workbook

        blocks: list[DocumentBlock] = []
        wb = load_workbook(str(path), data_only=True, read_only=False)
        try:
            for ws in wb.worksheets:
                if ws.max_row is None or ws.max_row == 0:
                    continue
                merged = _expand_merged_cells(ws)
                rows = _ws_rows(ws, merged)
                blocks.extend(self._rows_to_blocks(rows, ws.title))
        finally:
            wb.close()
        return blocks

    def _rows_to_blocks(self, rows: list[list], sheet_name: str) -> list[DocumentBlock]:
        # 表头 = 首个有内容行；空 sheet 无有效行则跳过
        try:
            header = next(row for row in rows if any(str(c).strip() for c in row))
        except StopIteration:
            return []
        header_idx = rows.index(header)
        data_rows = rows[header_idx + 1 :]
        blocks: list[DocumentBlock] = []
        for start, group in iter_row_groups(data_rows):
            # 行号从 1 起：表头占 header_idx 行，组内第 start+1..start+len 行是数据
            row_start = header_idx + start + 1
            row_end = header_idx + start + len(group)
            blocks.append(
                self._block(
                    table_text(header, group),
                    content_type="table",
                    heading_path=[sheet_name, f"行 {row_start}-{row_end}"],
                    metadata={
                        "sheet_name": sheet_name,
                        "row_start": row_start,
                        "row_end": row_end,
                        "table": rows_to_nested(group),
                        "table_format": "xlsx",
                    },
                )
            )
        return blocks


def _expand_merged_cells(ws) -> dict:
    """合并单元格区域：左上角值广播到区域内所有格（坐标 -> 值）。"""
    merged = {}
    for rng in ws.merged_cells.ranges:
        value = ws.cell(row=rng.min_row, column=rng.min_col).value
        if rng.min_row == rng.max_row and rng.min_col == rng.max_col:
            continue
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                if (row, col) != (rng.min_row, rng.min_col):
                    merged[(row, col)] = value
    return merged


def _ws_rows(ws, merged: dict) -> list[list]:
    """读取整个 sheet 为嵌套行；合并区域补值，None/空串规整为 ""。"""
    rows = []
    for row in ws.iter_rows():
        values = []
        for cell in row:
            if cell.value is None and (cell.row, cell.column) in merged:
                values.append(merged[(cell.row, cell.column)])
            else:
                values.append("" if cell.value is None else cell.value)
        if any(str(v).strip() for v in values):
            rows.append(values)
    return rows
