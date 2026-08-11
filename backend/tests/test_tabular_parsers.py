"""阶段5测试：xlsx/csv 解析器（合并单元格展开、行组分块、表头重复、profile 选型）。"""

import csv

from backend.app.rag.chunking_profiles import PROFILES, resolve_profile
from backend.app.rag.parsers.csv_parser import CsvParser
from backend.app.rag.parsers.xlsx_parser import XlsxParser
from backend.app.rag.parsers.tabular import iter_row_groups, table_text


def _write_xlsx(path, sheets):
    """sheets: {name: [[cell, ...], ...]}"""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    wb.save(path)
    return path


# ---------------------------------------------------------------- xlsx

def test_xlsx_single_sheet_becomes_table_blocks(tmp_path):
    path = _write_xlsx(
        tmp_path / "t.xlsx",
        {
            "Sheet1": [
                ["名称", "数量"],
                ["苹果", 3],
                ["香蕉", 5],
            ]
        },
    )
    blocks = XlsxParser("doc_x").parse(path)
    assert len(blocks) == 1
    block = blocks[0]
    assert block.content_type == "table"
    # 每行「列名：值」自包含
    assert "名称：苹果；数量：3" in block.text
    assert "名称：香蕉；数量：5" in block.text
    # heading_path 带 sheet 与行范围
    assert block.heading_path == ["Sheet1", "行 1-2"]
    assert block.metadata["sheet_name"] == "Sheet1"
    assert block.metadata["row_start"] == 1
    assert block.metadata["row_end"] == 2
    assert block.metadata["table"] == [["苹果", "3"], ["香蕉", "5"]]


def test_xlsx_merged_cells_expanded(tmp_path):
    # 数据区合并单元格：A3:C3 合并为"显示器专区"，值应广播到三列，不留空 cell
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["产品", "价格", "产地"])
    ws.append(["电脑", 5999, "北京"])
    ws.append(["显示器专区", None, None])
    ws.merge_cells("A3:C3")
    wb.save(tmp_path / "m.xlsx")

    blocks = XlsxParser("doc_m").parse(tmp_path / "m.xlsx")
    assert len(blocks) == 1
    block = blocks[0]
    # 合并区展开：值广播到整行，无空 cell
    assert block.metadata["table"][1] == ["显示器专区", "显示器专区", "显示器专区"]
    assert "产品：显示器专区；价格：显示器专区；产地：显示器专区" in block.text
    # 常规行不受影响
    assert "产品：电脑；价格：5999；产地：北京" in block.text


def test_xlsx_chunk_rows_slices_groups(tmp_path):
    rows = [["名称", "值"]] + [[f"项{i}", i] for i in range(1, 60)]
    path = _write_xlsx(tmp_path / "big.xlsx", {"D": rows})
    blocks = XlsxParser("doc_big").parse(path)
    # 59 数据行 + 表头，50 行一组 -> 2 块
    assert len(blocks) == 2
    assert blocks[0].heading_path == ["D", "行 1-50"]
    assert blocks[1].heading_path == ["D", "行 51-59"]
    # 组1 = 项1-50，组2 = 项51-59
    assert "名称：项1；值：1" in blocks[0].text
    assert "名称：项50；值：50" in blocks[0].text
    assert "名称：项51；值：51" in blocks[1].text
    assert "名称：项59；值：59" in blocks[1].text


def test_xlsx_empty_sheet_skipped(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("空表")
    wb.create_sheet("有数据")
    ws = wb["有数据"]
    ws.append(["a", "b"])
    ws.append([1, 2])
    wb.save(tmp_path / "empty.xlsx")

    blocks = XlsxParser("doc_e").parse(tmp_path / "empty.xlsx")
    assert len(blocks) == 1
    assert blocks[0].metadata["sheet_name"] == "有数据"


def test_xlsx_multiple_sheets_each_get_blocks(tmp_path):
    path = _write_xlsx(
        tmp_path / "multi.xlsx",
        {
            "A表": [["x", "y"], [1, 2]],
            "B表": [["m", "n"], [3, 4]],
        },
    )
    blocks = XlsxParser("doc_multi").parse(path)
    assert len(blocks) == 2
    assert [b.metadata["sheet_name"] for b in blocks] == ["A表", "B表"]


# ---------------------------------------------------------------- csv

def test_csv_becomes_table_blocks(tmp_path):
    p = tmp_path / "t.csv"
    with open(p, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["名称", "数量"])
        writer.writerow(["苹果", 3])
        writer.writerow(["香蕉", 5])
    blocks = CsvParser("doc_c").parse(p)
    assert len(blocks) == 1
    assert blocks[0].heading_path == ["t", "行 1-2"]
    assert "名称：苹果；数量：3" in blocks[0].text
    assert blocks[0].metadata["table_format"] == "csv"


def test_csv_with_bom_handled(tmp_path):
    p = tmp_path / "bom.csv"
    p.write_bytes(b"\xef\xbb\xbf" + "name,val\nfoo,1\n".encode("utf-8"))
    blocks = CsvParser("doc_bom").parse(p)
    assert len(blocks) == 1
    assert "name：foo；val：1" in blocks[0].text


# ---------------------------------------------------------------- 共享逻辑 + profile

def test_iter_row_groups_slices():
    rows = list(range(1, 121))
    groups = list(iter_row_groups(rows, chunk_rows=50))
    assert len(groups) == 3
    assert groups[0] == (0, list(range(1, 51)))
    assert groups[2] == (100, list(range(101, 121)))


def test_table_text_repeats_header_in_each_block():
    text = table_text(["名称", "数量"], [["苹果", 3], ["香蕉", 5]])
    assert "名称：苹果；数量：3" in text
    assert "名称：香蕉；数量：5" in text


def test_resolve_profile_picks_spreadsheet(tmp_path):
    path = _write_xlsx(tmp_path / "p.xlsx", {"S": [["a", "b"], [1, 2]]})
    blocks = XlsxParser("doc_p").parse(path)
    profile = resolve_profile("auto", "p.xlsx", blocks)
    assert profile.id == "spreadsheet"
    assert profile.chunker == "preserve"
    # 显式指定仍可用
    assert resolve_profile("technical", "p.xlsx", blocks).id == "technical"
    assert PROFILES["spreadsheet"].label == "表格数据"
